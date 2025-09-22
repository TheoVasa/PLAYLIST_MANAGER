import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(filename=self.file_path)
    
    def read_dataframe(self, sheet_name):
        """Lit une feuille Excel et retourne un DataFrame (structure conservée)."""
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine="openpyxl")
        return df

    def get_playlist_names(self):
        """Retourne la liste des playlists dans la feuille 'NOM PLAYLISTS'."""
        sheet_name = self.workbook.sheetnames[1]  # deuxième feuille
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine="openpyxl")
        return df['Playlists'].dropna().unique().tolist()

    def update_row(self, sheet_name, row_index, updates: dict):
        """
        Met à jour une ligne existante (en base 0 pour row_index) dans une feuille,
        sans casser les tables Excel.
        """
        sheet = self.workbook[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        for col_name, value in updates.items():
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                sheet.cell(row=row_index + 2, column=col_idx, value=value)  # +2 car Excel commence à 1 et +1 pour header

    def append_row(self, sheet_name, new_data: dict):
        """
        Ajoute une nouvelle ligne à la fin du tableau existant, en préservant la structure.
        """
        sheet = self.workbook[sheet_name]
        table = next((obj for obj in sheet._tables.values()), None)
        if not table:
            raise ValueError("Aucune table structurée trouvée sur cette feuille.")

        headers = [cell.value for cell in sheet[1]]
        col_index_map = {header: idx + 1 for idx, header in enumerate(headers)}

        # Calcul de la nouvelle ligne
        new_row_idx = sheet.max_row + 1

        # Ajouter les valeurs une par une avec styles copiés de la ligne précédente
        for header, value in new_data.items():
            col_idx = col_index_map.get(header)
            if col_idx:
                ref_cell = sheet.cell(row=new_row_idx - 1, column=col_idx)
                new_cell = sheet.cell(row=new_row_idx, column=col_idx)
                new_cell.value = value
                # Copier les styles de façon sécurisée
                new_cell.font = copy(ref_cell.font)
                new_cell.fill = copy(ref_cell.fill)
                new_cell.border = copy(ref_cell.border)
                new_cell.alignment = copy(ref_cell.alignment)
                new_cell.number_format = ref_cell.number_format

        # Étendre la table Excel
        end_col = sheet.max_column
        table.ref = f"A1:{get_column_letter(end_col)}{new_row_idx}"

        if "PLAYLIST" not in headers:
            raise ValueError(f"Colonne 'PLAYLIST' introuvable dans la feuille '{sheet_name}'.")

        col_idx = headers.index("PLAYLIST") + 1
        col_letter = get_column_letter(col_idx)

        # Supprimer les validations existantes pour cette colonne uniquement
        to_remove = []
        for dv in sheet.data_validations.dataValidation:
            for rng in dv.ranges.ranges:
                if CellRange(rng.coord).min_col == col_idx:
                    to_remove.append(dv)
                    break
        for dv in to_remove:
            sheet.data_validations.dataValidation.remove(dv)

        # Ajouter la nouvelle validation
        dv = DataValidation(
            type="list",
            formula1="=Playlists",
            allow_blank=True,
            showDropDown=False
        )
        dv.ranges.add(f"{col_letter}2:{col_letter}{new_row_idx}")
        sheet.add_data_validation(dv)

    def save(self):
        self.workbook.save(self.file_path)


